import csv
import argparse
from collections import OrderedDict
import os
from openpyxl import load_workbook

BLENDED_WEIGHTS = {"SP500": 0.40, "NASDAQ": 0.40, "RUSSELL": 0.20}
MAX_HOLDINGS = 100
MAX_POSITION = 0.08  # 8%


def normalize_text(value: str) -> str:
    return value.strip().lower().replace("_", " ")

def resolve_columns(fieldnames):
    lowered = {normalize_text(name): name for name in fieldnames or []}

    ticker_col = None
    market_cap_col = None
    name_col = None

    for key, original in lowered.items():

        if not ticker_col and (
            "ticker" in key
            or "symbol" in key
            or "security" in key
            or "holding" in key
            or "constituent" in key
        ):
            ticker_col = original

        if not market_cap_col and (
            "cap" in key
            or "market" in key
            or "weight" in key
            or "mkt" in key
            or "value" in key
            or "%" in key
        ):
            market_cap_col = original

        if not name_col and key in {"name", "company", "company name"}:
            name_col = original

    return ticker_col, market_cap_col, name_col

def parse_market_cap(value):
    if value is None:
        return 0.0

    text = str(value).strip().lower()
    text = text.replace("$", "").replace(",", "").replace(" ", "")

    if not text:
        return 0.0

    # Accept percentage-style values like 7.2%, 0.072%, or 7.2 percent.
    # These are treated as relative ranking values, not literal market caps.
    if text.endswith("%"):
        try:
            return float(text[:-1])
        except Exception:
            return 0.0

    if "percent" in text:
        try:
            return float(text.replace("percent", ""))
        except Exception:
            return 0.0

    multiplier = 1.0

    if "trillion" in text:
        multiplier = 1_000_000_000_000
        text = text.replace("trillion", "")
    elif "billion" in text:
        multiplier = 1_000_000_000
        text = text.replace("billion", "")
    elif "million" in text:
        multiplier = 1_000_000
        text = text.replace("million", "")
    elif text.endswith("t"):
        multiplier = 1_000_000_000_000
        text = text[:-1]
    elif text.endswith("b"):
        multiplier = 1_000_000_000
        text = text[:-1]
    elif text.endswith("m"):
        multiplier = 1_000_000
        text = text[:-1]

    try:
        return float(text) * multiplier
    except Exception:
        return 0.0


def read_index(path, source_name):
    rows = []

    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        wb = load_workbook(path, data_only=True)
        ws = wb.active or wb.worksheets[0]

        values = list(ws.iter_rows(values_only=True))
        if not values:
            raise ValueError(f"{source_name}: empty Excel file")

        headers = [str(h).strip() if h is not None else "" for h in values[0]]
        ticker_col, market_cap_col, name_col = resolve_columns(headers)

        if not ticker_col:
            raise ValueError(f"{source_name}: could not find ticker column")

        if not market_cap_col:
            raise ValueError(f"{source_name}: could not find market cap/weight column")

        ticker_idx = headers.index(ticker_col)
        market_cap_idx = headers.index(market_cap_col)
        name_idx = headers.index(name_col) if name_col else None

        for row in values[1:]:
            ticker = str(row[ticker_idx]).strip() if row[ticker_idx] is not None else ""
            if not ticker:
                continue

            mc = parse_market_cap(row[market_cap_idx])

            rows.append({
                "ticker": ticker,
                "name": str(row[name_idx]).strip() if name_idx is not None and row[name_idx] is not None else "",
                "market_cap": mc,
                "source": source_name,
            })

    else:
        encodings = ["utf-8", "utf-8-sig", "cp1252", "latin1"]
        last_error = None
        reader = None
        f = None

        for encoding in encodings:
            try:
                f = open(path, newline="", encoding=encoding)
                reader = csv.DictReader(f)
                break
            except Exception as e:
                last_error = e
                if f:
                    f.close()

        if reader is None:
            raise ValueError(f"Could not read CSV: {last_error}")

        try:
            ticker_col, market_cap_col, name_col = resolve_columns(reader.fieldnames)

            if not ticker_col:
                raise ValueError(f"{source_name}: could not find ticker column")

            if not market_cap_col:
                raise ValueError(f"{source_name}: could not find market cap/weight column")

            for r in reader:
                ticker = (r.get(ticker_col) or "").strip()
                if not ticker:
                    continue

                mc = parse_market_cap(r.get(market_cap_col))

                rows.append({
                    "ticker": ticker,
                    "name": (r.get(name_col) or "").strip() if name_col else "",
                    "market_cap": mc,
                    "source": source_name,
                })
        finally:
            if f:
                f.close()

    rows.sort(key=lambda r: r["market_cap"], reverse=True)
    return rows

def select_from_index(index_rows, target_weight, already_selected, max_total_slots):
    # Select constituents excluding already_selected, up to available slots
    selected = []
    for r in index_rows:
        if r["ticker"] in already_selected:
            continue
        selected.append(r)
        if len(selected) >= max_total_slots:
            break
    # assign equal weight within this slice
    if not selected:
        return []
    per = target_weight / len(selected)
    return [{**r, "target_weight": per} for r in selected]


def enforce_caps(holdings):
    # holdings: list of dicts with ticker and target_weight (fractions)
    # enforce MAX_POSITION cap by iterative redistribution
    # convert to OrderedDict keyed by ticker
    hd = OrderedDict()
    for h in holdings:
        hd[h["ticker"]] = {**h}

    # iterative caps
    while True:
        over = {t: v for t, v in hd.items() if v["target_weight"] > MAX_POSITION}
        if not over:
            break
        excess = 0.0
        for t, v in over.items():
            excess += v["target_weight"] - MAX_POSITION
            v["target_weight"] = MAX_POSITION
            hd[t] = v
        # redistribute excess to unconstrained holdings proportionally
        unconstrained = {t: v for t, v in hd.items() if v["target_weight"] < MAX_POSITION}
        if not unconstrained:
            # nothing to redistribute
            break
        total_un = sum(v["target_weight"] for v in unconstrained.values())
        if total_un <= 0:
            # spread equally
            per = excess / len(unconstrained)
            for t in unconstrained:
                hd[t]["target_weight"] += per
        else:
            for t in unconstrained:
                add = excess * (hd[t]["target_weight"] / total_un)
                hd[t]["target_weight"] += add
    # normalize to sum to 1
    total = sum(v["target_weight"] for v in hd.values())
    if total <= 0:
        return list(hd.values())
    for v in hd.values():
        v["target_weight"] = v["target_weight"] / total
    return list(hd.values())


def build_portfolio(sp500_csv, nasdaq100_csv, russell_csv, out_csv):
    sp = read_index(sp500_csv, "SP500")
    nd = read_index(nasdaq100_csv, "NASDAQ")
    ru = read_index(russell_csv, "RUSSELL")

    selected = []
    tickers = set()

    # Heuristic for slots per slice: split MAX_HOLDINGS proportionally to blended weights
    slots_sp = max(1, int(round(MAX_HOLDINGS * BLENDED_WEIGHTS["SP500"])))
    slots_nd = max(1, int(round(MAX_HOLDINGS * BLENDED_WEIGHTS["NASDAQ"])))
    slots_ru = MAX_HOLDINGS - slots_sp - slots_nd
    if slots_ru < 1:
        slots_ru = 1

    # Nasdaq slice: select from Nasdaq-100 only
    nd_sel = select_from_index(nd, BLENDED_WEIGHTS["NASDAQ"], tickers, slots_nd)
    for h in nd_sel:
        tickers.add(h["ticker"])
        selected.append(h)

    # S&P500 slice
    sp_sel = select_from_index(sp, BLENDED_WEIGHTS["SP500"], tickers, slots_sp)
    for h in sp_sel:
        tickers.add(h["ticker"])
        selected.append(h)

    # Russell slice (exclude overlaps)
    ru_sel = select_from_index(ru, BLENDED_WEIGHTS["RUSSELL"], tickers, slots_ru)
    for h in ru_sel:
        tickers.add(h["ticker"])
        selected.append(h)

    # If we have fewer than MAX_HOLDINGS, consider adding more from Russell by market cap
    if len(selected) < MAX_HOLDINGS:
        needed = MAX_HOLDINGS - len(selected)
        for r in ru:
            if r["ticker"] in tickers:
                continue
            selected.append({**r, "target_weight": 0.0})
            tickers.add(r["ticker"])
            needed -= 1
            if needed <= 0:
                break

    # If more than MAX_HOLDINGS, trim by market cap
    if len(selected) > MAX_HOLDINGS:
        selected.sort(key=lambda x: x.get("market_cap", 0), reverse=True)
        selected = selected[:MAX_HOLDINGS]

    # For any holdings without assigned target_weight (added to fill slots), assign a small share
    # Start by giving each slice-assigned holding its per-slice equal weight; those that were added as fillers get a tiny placeholder
    # We'll now ensure the holdings have weights summing to 1 and enforce caps
    total_assigned = sum(h.get("target_weight", 0) for h in selected)
    unassigned = [h for h in selected if h.get("target_weight", 0) == 0]
    if unassigned:
        # give them a tiny equal share of remaining weight
        remaining = max(0.0, 1.0 - total_assigned)
        per = remaining / len(unassigned)
        for h in unassigned:
            h["target_weight"] = per

    # enforce caps and normalize
    final = enforce_caps(selected)

    # write out CSV
    with open(out_csv, "w", newline='', encoding='utf-8') as f:
        fieldnames = ["ticker", "name", "market_cap", "source", "weight_pct"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for h in final:
            writer.writerow({
                "ticker": h["ticker"],
                "name": h.get("name", ""),
                "market_cap": "{:.2f}".format(h.get("market_cap", 0.0)),
                "source": h.get("source", ""),
                "weight_pct": "{:.4f}".format(h.get("target_weight", 0.0) * 100),
            })


def main():
    p = argparse.ArgumentParser(description="Build a non-taxable blended-index portfolio")
    p.add_argument("--sp500", required=True, help="CSV of S&P 500 constituents")
    p.add_argument("--nasdaq100", required=True, help="CSV of Nasdaq-100 constituents")
    p.add_argument("--russell", required=True, help="CSV of Russell 3000 constituents")
    p.add_argument("--out", default="portfolio.csv", help="Output CSV file")
    args = p.parse_args()
    build_portfolio(args.sp500, args.nasdaq100, args.russell, args.out)
    print(f"Portfolio written to {args.out}")


if __name__ == "__main__":
    main()
